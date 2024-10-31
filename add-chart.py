from openpyxl import load_workbook # Import the load_workbook function to work with Excel files
from openpyxl.chart import BarChart, Reference # Import BarChart and Reference classes for creating charts

# Load the Excel workbook named 'pivot_table.xlsx'
wb = load_workbook('pivot_table.xlsx')
# Select the sheet named 'Report' from the workbook
sheet = wb ['Report']

# Get the minimum and maximum column indices from the active sheet
min_column = wb.active.min_column
max_column = wb.active.max_column
# Get the minimum and maximum row indices from the active sheet
min_row = wb.active.min_row
max_row = wb.active.max_row

# Create a new BarChart object
barchart = BarChart()
# Create a Reference object for the data to be plotted (excluding the first column for labels)
data = Reference(sheet, min_col = min_column + 1, max_col = max_column, min_row = min_row, max_row = max_row)
# Create a Reference object for the categories (using the first column for labels)
categories   = Reference(sheet, min_col = min_column, max_col = min_column, min_row = min_row + 1, max_row = max_row )
# Add the data to the bar chart, using the first row of data as titles
barchart.add_data(data, titles_from_data=True)
# Set the categories for the bar chart
barchart.set_categories(categories)
# Add the bar chart to the sheet at the specified cell location ("B12")
sheet.add_chart(barchart, "B12")
# Set the title of the bar chart
barchart.title = "Sales by Product line"
# Set the style of the bar chart (style 5 is a predefined style in openpyxl)
barchart.style = 5
# Save the modified workbook with the new chart as 'Barchart.xlsx'
wb.save("Barchart.xlsx")
